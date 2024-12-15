from django.utils.decorators import method_decorator
from django.utils.timezone import get_current_timezone
from django.views.decorators.csrf import csrf_exempt
from django.views import View
from django.http import JsonResponse
from django.http import HttpResponse
from django.db.models import Count
from robots.models import Robot, RobotModel
from openpyxl.styles import Alignment
from openpyxl import Workbook
import datetime
import json




@method_decorator(csrf_exempt, name='dispatch')
class RobotAPIView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            model = data.get('model')
            version = data.get('version')
            created = data.get('created')

            # Validate required fields
            if not all([model, version, created]):
                return JsonResponse({"error": "All fields are required."}, status=400)

            # Validate model existence
            if not RobotModel.objects.filter(model=model).exists():
                return JsonResponse({"error": "Model does not exist."}, status=400)
            
            created_dt = datetime.datetime.fromisoformat(created)
            current_timezone = get_current_timezone()
            created_aware = created_dt.replace(tzinfo=current_timezone)

            # Create and save a new Robot instance
            robot = Robot.objects.create(
                model=model,
                version=version,
                created=created_aware
            )

            return JsonResponse({"message": "Robot created successfully.", "serial": robot.serial}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format.", "example": {"model":"A1","version":"V2","created":"2022-12-31 23:59:59"}}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


def generate_excel_report():
    # Creating an Excel file
    wb = Workbook()

    # We get the start and end dates of the last week
    today = datetime.date.today()
    start_of_week = today - datetime.timedelta(days=today.weekday() + 7)
    end_of_week = start_of_week + datetime.timedelta(days=6)

    # Getting data from the database
    robots = (
        Robot.objects.filter(created__range=[start_of_week, end_of_week])
        .values("model", "version")
        .annotate(total=Count("id"))
    )

    if not robots:
        # If there is no data, add a standard sheet with a message
        ws = wb.active
        ws.title = "Нет данных"
        ws.append(["На этой неделе роботы не производились."])
        ws.column_dimensions["A"].width = 50  # Increasing the width for the message
        for cell in ws[1]:  # Align the message
            cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        # Creating pages for each model
        models = set(robot["model"] for robot in robots)

        for model in models:
            ws = wb.create_sheet(title=model)

            # Adding headlines
            ws.append(["Модель", "Версия", "Кол-во за неделю"])
            
            # Setting the column width
            ws.column_dimensions["A"].width = 10  # Width of the "Model" column
            ws.column_dimensions["B"].width = 10  # The width of the "Version" column
            ws.column_dimensions["C"].width = 20  # The width of the column "Quantity per week"

            # Align the headings in the center
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Filtering data by the current model
            model_data = [robot for robot in robots if robot["model"] == model]
            
            # Adding data for each version of the model
            for robot in model_data:
                row = [robot["model"], robot["version"], robot["total"]]
                ws.append(row)

            # Align the data to the center
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # We delete the standard sheet if it is not needed
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    return wb


def download_excel_report(request):
    # Generating an Excel file
    wb = generate_excel_report()
    
    # Configuring the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="robots_report.xlsx"'
    
    # Saving the file in response
    wb.save(response)
    return response